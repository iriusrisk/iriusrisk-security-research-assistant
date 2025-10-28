"""
XLSX export service for IriusRisk Library Editor API
"""

import logging
from typing import Dict, List, Set
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook

from isra.src.ile.backend.app.configuration import ExcelConstants
from isra.src.ile.backend.app.models import (
    ILEVersion, IRCategoryComponent, IRComponentDefinition, IRControl,
    IRLibrary, IRReference, IRRelation, IRRiskPattern, IRRule,
    IRRuleAction, IRRuleCondition, IRStandard, IRSupportedStandard,
    IRTest, IRThreat, IRUseCase, IRWeakness
)
from isra.src.ile.backend.app.services.data_service import DataService

logger = logging.getLogger(__name__)


class XLSXExportService:
    """Service for exporting libraries to XLSX format"""
    
    def __init__(self):
        self.data_service = DataService()
    
    def export_library_xlsx(self, lib: IRLibrary, version: ILEVersion, version_path: str) -> None:
        """Export library to XLSX file"""
        xlsx_file_path = Path(version_path) / lib.filename.replace("xml", "xlsx")
        
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create sheets
            risk_patterns_sheet = workbook.create_sheet("Risk Patterns")
            usecases_sheet = workbook.create_sheet("Use Cases")
            threats_sheet = workbook.create_sheet("Threats")
            weaknesses_sheet = workbook.create_sheet("Weaknesses")
            controls_sheet = workbook.create_sheet("Controls")
            references_sheet = workbook.create_sheet("References")
            relations_sheet = workbook.create_sheet("Relations")
            rules_sheet = workbook.create_sheet("Rules")
            properties_sheet = workbook.create_sheet("Library properties")
            components_sheet = workbook.create_sheet("Components")
            standards_sheet = workbook.create_sheet("Standards")
            supported_standards_sheet = workbook.create_sheet("Supported standards")
            
            # Create sheet content
            self._create_risk_patterns_sheet(risk_patterns_sheet, lib)
            self._create_rules_sheet(rules_sheet, lib)
            self._create_library_properties_sheet(properties_sheet, lib)
            self._create_relations_sheet(relations_sheet, lib)
            self._create_references_sheet(references_sheet, lib, version)
            self._create_usecases_sheet(usecases_sheet, lib, version)
            self._create_threats_sheet(threats_sheet, lib, version)
            self._create_weaknesses_sheet(weaknesses_sheet, lib, version)
            self._create_controls_sheet(controls_sheet, lib, version)
            self._create_components_sheet(components_sheet, lib, version)
            self._create_standards_sheet(standards_sheet, lib, version)
            self._create_supported_standards_sheet(supported_standards_sheet, lib, version)
            
            # Save workbook
            workbook.save(xlsx_file_path)
            logger.info(f"Export process finished: {xlsx_file_path}")
            
        except Exception as e:
            logger.error(f"Exception while saving {lib.ref} to XLSX: {e}")
            raise RuntimeError(f"Failed to export library to XLSX: {e}") from e
    
    def _create_references_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create references sheet"""
        # Headers
        self._set_value_and_color(sheet, 1, 1, "Name", ExcelConstants.RISK_PATTERN_HEADER, True)
        self._set_value_and_color(sheet, 1, 2, "URL", ExcelConstants.RISK_PATTERN_HEADER, True)
        self._set_value_and_color(sheet, 1, 3, "UUID", ExcelConstants.RISK_PATTERN_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        for reference in version.references.values():
            color = ExcelConstants.RISK_PATTERN_COLOR_1 if alternating_color else ExcelConstants.RISK_PATTERN_COLOR_2
            self._set_value_and_color(sheet, working_row, 1, reference.name, color, False)
            self._set_value_and_color(sheet, working_row, 2, reference.url, color, False)
            self._set_value_and_color(sheet, working_row, 3, reference.uuid, color, False)
            
            alternating_color = not alternating_color
            working_row += 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 100.0
        sheet.column_dimensions['B'].width = 100.0
    
    def _create_risk_patterns_sheet(self, sheet, lib: IRLibrary) -> None:
        """Create risk patterns sheet"""
        # Headers
        self._set_value_and_color(sheet, 1, 1, "Ref", ExcelConstants.RISK_PATTERN_HEADER, True)
        self._set_value_and_color(sheet, 1, 2, "Name", ExcelConstants.RISK_PATTERN_HEADER, True)
        self._set_value_and_color(sheet, 1, 3, "Desc", ExcelConstants.RISK_PATTERN_HEADER, True)
        self._set_value_and_color(sheet, 1, 4, "UUID", ExcelConstants.RISK_PATTERN_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        for rp in lib.risk_patterns.values():
            color = ExcelConstants.RISK_PATTERN_COLOR_1 if alternating_color else ExcelConstants.RISK_PATTERN_COLOR_2
            self._set_value_and_color(sheet, working_row, 1, rp.ref, color, False)
            self._set_value_and_color(sheet, working_row, 2, rp.name, color, False)
            self._set_value_and_color(sheet, working_row, 3, rp.desc, color, False)
            self._set_value_and_color(sheet, working_row, 4, rp.uuid, color, False)
            
            alternating_color = not alternating_color
            working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_usecases_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create use cases sheet"""
        # Headers
        self._set_value_and_color(sheet, 1, 1, "Ref", ExcelConstants.USE_CASE_HEADER, True)
        self._set_value_and_color(sheet, 1, 2, "Name", ExcelConstants.USE_CASE_HEADER, True)
        self._set_value_and_color(sheet, 1, 3, "Desc", ExcelConstants.USE_CASE_HEADER, True)
        self._set_value_and_color(sheet, 1, 4, "UUID", ExcelConstants.USE_CASE_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        usecases = self._get_list_from_relations(lib, "usecases")
        
        for uc_ref in usecases:
            usecase = version.usecases.get(uc_ref)
            if usecase:
                color = ExcelConstants.USE_CASE_COLOR_1 if alternating_color else ExcelConstants.USE_CASE_COLOR_2
                self._set_value_and_color(sheet, working_row, 1, usecase.ref, color, False)
                self._set_value_and_color(sheet, working_row, 2, usecase.name, color, False)
                self._set_value_and_color(sheet, working_row, 3, usecase.desc, color, False)
                self._set_value_and_color(sheet, working_row, 4, usecase.uuid, color, False)
                
                alternating_color = not alternating_color
                working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_threats_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create threats sheet"""
        # Headers
        headers = [
            "Ref", "Name", "Desc", "Confidentiality", "Integrity", "Availability",
            "Ease Of Exploitation", "References", "Mitre", "STRIDE", "UUID"
        ]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.THREAT_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        threats = self._get_list_from_relations(lib, "threats")
        
        for threat_ref in threats:
            threat = version.threats.get(threat_ref)
            if threat:
                # Format references
                threat_references = [f"{key}:{value}" for key, value in threat.references.items()]
                refs = ExcelConstants.SEPARATOR.join(threat_references)
                
                color = ExcelConstants.THREAT_COLOR_1 if alternating_color else ExcelConstants.THREAT_COLOR_2
                values = [
                    threat.ref, threat.name, threat.desc,
                    threat.risk_rating.confidentiality if threat.risk_rating else "",
                    threat.risk_rating.integrity if threat.risk_rating else "",
                    threat.risk_rating.availability if threat.risk_rating else "",
                    threat.risk_rating.ease_of_exploitation if threat.risk_rating else "",
                    refs,
                    ExcelConstants.SEPARATOR.join(threat.mitre or []),
                    ExcelConstants.SEPARATOR.join(threat.stride or []),
                    threat.uuid
                ]
                
                for i, value in enumerate(values, 1):
                    self._set_value_and_color(sheet, working_row, i, value, color, False)
                
                alternating_color = not alternating_color
                working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_weaknesses_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create weaknesses sheet"""
        # Headers
        headers = ["Ref", "Name", "Desc", "Impact", "Test Steps", "Test References", "UUID"]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.WEAKNESS_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        weaknesses = self._get_list_from_relations(lib, "weaknesses")
        
        for weakness_ref in weaknesses:
            weakness = version.weaknesses.get(weakness_ref)
            if weakness:
                # Format test references
                test_references = [f"{key}:{value}" for key, value in weakness.test.references.items()]
                t_refs = ExcelConstants.SEPARATOR.join(test_references)
                
                color = ExcelConstants.WEAKNESS_COLOR_1 if alternating_color else ExcelConstants.WEAKNESS_COLOR_2
                values = [
                    weakness.ref, weakness.name, weakness.desc, weakness.impact,
                    weakness.test.steps, t_refs, weakness.uuid
                ]
                
                for i, value in enumerate(values, 1):
                    self._set_value_and_color(sheet, working_row, i, value, color, False)
                
                alternating_color = not alternating_color
                working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_controls_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create controls sheet"""
        # Headers
        headers = [
            "Ref", "Name", "Desc", "State", "Cost", "References", "Test Steps",
            "Test References", "Standards", "Implementations", "Base Standard",
            "Base Standard Section", "Scope", "MITRE", "UUID"
        ]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.COUNTERMEASURE_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        controls = self._get_list_from_relations(lib, "controls")
        
        for control_ref in controls:
            control = version.controls.get(control_ref)
            if control:
                # Format references
                control_references = [f"{key}:{value}" for key, value in control.references.items()]
                refs = ExcelConstants.SEPARATOR.join(control_references)
                
                # Format test references
                test_references = [f"{key}:{value}" for key, value in control.test.references.items()]
                t_refs = ExcelConstants.SEPARATOR.join(test_references)
                
                # Format standards
                standard_list = [f"{key}:{value}" for key, value in control.standards.items()]
                sts = ExcelConstants.SEPARATOR.join(standard_list)
                
                # Format implementations
                impl = ExcelConstants.SEPARATOR.join(control.implementations)
                
                color = ExcelConstants.COUNTERMEASURE_COLOR_1 if alternating_color else ExcelConstants.COUNTERMEASURE_COLOR_2
                values = [
                    control.ref, control.name, control.desc, control.state, control.cost,
                    refs, control.test.steps, t_refs, sts, impl,
                    ExcelConstants.SEPARATOR.join(control.base_standard or []),
                    ExcelConstants.SEPARATOR.join(control.base_standard_section or []),
                    ExcelConstants.SEPARATOR.join(control.scope or []),
                    ExcelConstants.SEPARATOR.join(control.mitre or []),
                    control.uuid
                ]
                
                for i, value in enumerate(values, 1):
                    self._set_value_and_color(sheet, working_row, i, value, color, False)
                
                alternating_color = not alternating_color
                working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_relations_sheet(self, sheet, lib: IRLibrary) -> None:
        """Create relations sheet"""
        # Headers
        headers = ["Risk Pattern", "Use Case", "Threat", "Weakness", "Control", "Mitigation"]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.RISK_PATTERN_HEADER, True)
        
        alternating_color = True
        working_row = 2
        
        for rel in lib.relations.values():
            color = ExcelConstants.RISK_PATTERN_COLOR_1 if alternating_color else ExcelConstants.RISK_PATTERN_COLOR_2
            values = [
                rel.risk_pattern_uuid, rel.usecase_uuid, rel.threat_uuid,
                rel.weakness_uuid, rel.control_uuid, rel.mitigation
            ]
            
            for i, value in enumerate(values, 1):
                self._set_value_and_color(sheet, working_row, i, value, color, False)
            
            alternating_color = not alternating_color
            working_row += 1
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_rules_sheet(self, sheet, lib: IRLibrary) -> None:
        """Create rules sheet"""
        # Headers
        self._set_rules_excel_headers(sheet)
        
        # Rules data
        rule_color = True
        condition_color = True
        action_color = True
        
        working_row = 2
        for rule in lib.rules:
            rule_working_row = working_row
            r_color = ExcelConstants.RULE_COLOR_1 if rule_color else ExcelConstants.RULE_COLOR_2
            
            self._set_value_and_color(sheet, working_row, ExcelConstants.RULES_NAME, rule.name, r_color, False)
            self._set_value_and_color(sheet, working_row, ExcelConstants.RULES_MODULE, rule.module, r_color, False)
            self._set_value_and_color(sheet, working_row, ExcelConstants.RULES_GUI, rule.generated_by_gui, r_color, False)
            
            # Conditions
            cond_working_row = rule_working_row
            for cond in rule.conditions:
                c_color = ExcelConstants.RULE_CONDITION_COLOR_1 if condition_color else ExcelConstants.RULE_CONDITION_COLOR_2
                self._set_value_and_color(sheet, cond_working_row, ExcelConstants.RULES_CONDITION_NAME, cond.name, c_color, False)
                self._set_value_and_color(sheet, cond_working_row, ExcelConstants.RULES_CONDITION_VALUE, cond.value, c_color, False)
                self._set_value_and_color(sheet, cond_working_row, ExcelConstants.RULES_CONDITION_FIELD, cond.field, c_color, False)
                condition_color = not condition_color
                cond_working_row += 1
            
            # Actions
            act_working_row = rule_working_row
            for act in rule.actions:
                a_color = ExcelConstants.RULE_ACTION_COLOR_1 if action_color else ExcelConstants.RULE_ACTION_COLOR_2
                self._set_value_and_color(sheet, act_working_row, ExcelConstants.RULES_ACTION_NAME, act.name, a_color, False)
                self._set_value_and_color(sheet, act_working_row, ExcelConstants.RULES_ACTION_VALUE, act.value, a_color, False)
                self._set_value_and_color(sheet, act_working_row, ExcelConstants.RULES_ACTION_PROJECT, act.project, a_color, False)
                action_color = not action_color
                act_working_row += 1
            
            # Merge cells for rule name, module, and GUI
            if rule.actions or rule.conditions:
                working_row = max(cond_working_row, act_working_row)
            
            working_row += 1
            rule_color = not rule_color
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RULES_LAST_HEADER)
    
    def _create_library_properties_sheet(self, sheet, lib: IRLibrary) -> None:
        """Create library properties sheet"""
        # Headers
        self._set_value_and_color(sheet, 1, 1, "General", ExcelConstants.LIBRARY_PROPERTY_HEADER, True)
        self._set_value_and_color(sheet, 1, 2, "Values", ExcelConstants.LIBRARY_PROPERTY_HEADER, True)
        
        # Library properties
        properties = [
            ("Library Name", lib.name),
            ("Library Ref", lib.ref),
            ("Library Desc", lib.desc),
            ("Revision", lib.revision),
            ("Enabled", lib.enabled)
        ]
        
        for i, (label, value) in enumerate(properties, 2):
            color1 = ExcelConstants.LIBRARY_PROPERTY_COLOR_1 if i % 2 == 0 else ExcelConstants.LIBRARY_PROPERTY_COLOR_2
            color2 = ExcelConstants.LIBRARY_PROPERTY_COLOR_1 if i % 2 == 0 else ExcelConstants.LIBRARY_PROPERTY_COLOR_2
            
            self._set_value_and_color(sheet, i, 1, label, color1, False)
            self._set_value_and_color(sheet, i, 2, value, color2, False)
        
        self._adjust_height_and_width(sheet, 7, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_components_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create components sheet"""
        # Headers
        headers = [
            "Component Definition Name", "Component Definition Ref", "Component Definition Desc",
            "Category Name", "Category Ref", "Category UUID", "Risk Patterns", "Visible", "Component UUID"
        ]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.LIBRARY_COMPONENT_HEADER, True)
        
        # Create category lookup
        categories_by_ref = {cat.ref: cat for cat in version.categories.values()}
        
        component_color = True
        working_row = 2
        
        for cd in lib.component_definitions.values():
            category = categories_by_ref.get(cd.category_ref)
            
            color = ExcelConstants.LIBRARY_COMPONENT_COLOR_1 if component_color else ExcelConstants.LIBRARY_COMPONENT_COLOR_2
            values = [
                cd.name, cd.ref, cd.desc,
                category.name if category else "",
                cd.category_ref,
                category.uuid if category else "",
                ",".join(cd.risk_pattern_refs),
                cd.visible, cd.uuid
            ]
            
            for i, value in enumerate(values, 1):
                self._set_value_and_color(sheet, working_row, i, value, color, False)
            
            working_row += 1
            component_color = not component_color
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_standards_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create standards sheet"""
        # Headers
        headers = ["Supported Standard Ref", "Standard Ref", "Standard UUID"]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.LIBRARY_STANDARD_HEADER, True)
        
        standard_color = True
        working_row = 2
        
        for standard in version.standards.values():
            color = ExcelConstants.LIBRARY_STANDARD_COLOR_1 if standard_color else ExcelConstants.LIBRARY_STANDARD_COLOR_2
            values = [standard.supported_standard_ref, standard.standard_ref, standard.uuid]
            
            for i, value in enumerate(values, 1):
                self._set_value_and_color(sheet, working_row, i, value, color, False)
            
            working_row += 1
            standard_color = not standard_color
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _create_supported_standards_sheet(self, sheet, lib: IRLibrary, version: ILEVersion) -> None:
        """Create supported standards sheet"""
        # Headers
        headers = ["Supported Standard Name", "Supported Standard Ref", "Supported Standard UUID"]
        for i, header in enumerate(headers, 1):
            self._set_value_and_color(sheet, 1, i, header, ExcelConstants.LIBRARY_STANDARD_HEADER, True)
        
        standard_color = True
        working_row = 2
        
        for supported_standard in version.supported_standards.values():
            color = ExcelConstants.LIBRARY_STANDARD_COLOR_1 if standard_color else ExcelConstants.LIBRARY_STANDARD_COLOR_2
            values = [supported_standard.name, supported_standard.ref, supported_standard.uuid]
            
            for i, value in enumerate(values, 1):
                self._set_value_and_color(sheet, working_row, i, value, color, False)
            
            working_row += 1
            standard_color = not standard_color
        
        self._adjust_height_and_width(sheet, working_row, ExcelConstants.RISK_PATTERNS_LAST_HEADER)
    
    def _set_rules_excel_headers(self, sheet) -> None:
        """Set rules Excel headers"""
        headers = [
            (ExcelConstants.RULES_NAME, "Rule Name", ExcelConstants.RULE_HEADER),
            (ExcelConstants.RULES_MODULE, "Module", ExcelConstants.RULE_HEADER),
            (ExcelConstants.RULES_GUI, "Generated by GUI", ExcelConstants.RULE_HEADER),
            (ExcelConstants.RULES_CONDITION_NAME, "Condition Name", ExcelConstants.RULE_CONDITION_HEADER),
            (ExcelConstants.RULES_CONDITION_VALUE, "Condition Value", ExcelConstants.RULE_CONDITION_HEADER),
            (ExcelConstants.RULES_CONDITION_FIELD, "Condition Field", ExcelConstants.RULE_CONDITION_HEADER),
            (ExcelConstants.RULES_ACTION_NAME, "Action Name", ExcelConstants.RULE_ACTION_HEADER),
            (ExcelConstants.RULES_ACTION_VALUE, "Action Value", ExcelConstants.RULE_ACTION_HEADER),
            (ExcelConstants.RULES_ACTION_PROJECT, "Action Project", ExcelConstants.RULE_ACTION_HEADER)
        ]
        
        for col, header, color in headers:
            self._set_value_and_color(sheet, 1, col, header, color, True)
    
    def _adjust_height_and_width(self, sheet, limit_row: int, limit_col: int) -> None:
        """Adjust height and width of Excel file"""
        for i in range(1, limit_col + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 30.0
        
        for i in range(1, limit_row + 1):
            sheet.row_dimensions[i].height = 15.0
    
    def _set_value_and_color(self, sheet, row: int, col: int, value: str, color: str, header: bool) -> None:
        """Set value and style at the same time"""
        cell = sheet.cell(row=row, column=col, value=value)
        
        # Create styles
        font = Font(bold=header, color="FFFFFF" if header else "000000")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=not header)
        border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="thin", color="FFFFFF")
        )
        
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
    
    def _get_list_from_relations(self, lib: IRLibrary, attrib: str) -> Set[str]:
        """Get list of elements from relations"""
        values_in_library = set()
        for rel in lib.relations.values():
            if attrib == "controls" and rel.control_uuid:
                values_in_library.add(rel.control_uuid)
            elif attrib == "weaknesses" and rel.weakness_uuid:
                values_in_library.add(rel.weakness_uuid)
            elif attrib == "threats" and rel.threat_uuid:
                values_in_library.add(rel.threat_uuid)
            elif attrib == "usecases" and rel.usecase_uuid:
                values_in_library.add(rel.usecase_uuid)
        
        return values_in_library
