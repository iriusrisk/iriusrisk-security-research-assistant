"""
XLSX import service for IriusRisk Library Editor API
"""

import logging
import uuid
from typing import BinaryIO, List, Optional

import openpyxl
from openpyxl.workbook import Workbook

from isra.src.ile.backend.app.configuration import ExcelConstants
from isra.src.ile.backend.app.models import (
    ILEVersion, IRCategoryComponent, IRComponentDefinition, IRControl,
    IRLibrary, IRReference, IRRelation, IRRiskPattern, IRRiskRating,
    IRRule, IRRuleAction, IRRuleCondition, IRStandard, IRSupportedStandard,
    IRTest, IRThreat, IRUseCase, IRWeakness
)

logger = logging.getLogger(__name__)


class XLSXImportService:
    """Service for importing XLSX library files"""
    
    def import_library_xlsx(self, filename: str, library: BinaryIO, version_element: ILEVersion) -> None:
        """Import library from XLSX stream"""
        try:
            filename = filename.replace("xlsx", "xml").replace("xls", "xml")
            
            # Load workbook
            workbook = openpyxl.load_workbook(library)
            
            # Library properties
            properties_sheet = workbook["Library properties"]
            library_name = properties_sheet.cell(2, 2).value or ""
            library_ref = properties_sheet.cell(3, 2).value or ""
            library_desc = properties_sheet.cell(4, 2).value or ""
            library_revision = properties_sheet.cell(5, 2).value or "1"
            library_enabled = properties_sheet.cell(6, 2).value or "true"
            
            if not library_enabled:
                library_enabled = "true"
            
            new_library = IRLibrary(
                ref=library_ref,
                name=library_name,
                desc=library_desc,
                revision=str(library_revision),
                filename=filename,
                enabled=str(library_enabled)
            )
            
            # Import various elements
            self._add_references_from_excel(version_element, workbook["References"])
            self._add_supported_standards_from_excel(version_element, workbook["Supported standards"])
            self._add_standards_from_excel(version_element, workbook["Standards"])
            self._add_category_components_from_excel(version_element, workbook["Components"])
            self._add_component_definitions_from_excel(new_library, workbook["Components"])
            self._add_risk_patterns_from_excel(new_library, workbook["Risk Patterns"])
            self._add_usecases_from_excel(version_element, workbook["Use Cases"])
            self._add_threats_from_excel(version_element, workbook["Threats"])
            self._add_weaknesses_from_excel(version_element, workbook["Weaknesses"])
            self._add_controls_from_excel(version_element, workbook["Controls"])
            self._add_relations_from_excel(new_library, workbook["Relations"])
            self._add_rules_from_excel(new_library, workbook["Rules"])
            
            logger.info(f"Adding new library {library_ref} to version {version_element.version}")
            version_element.libraries[library_ref] = new_library
            
        except Exception as e:
            logger.error(f"Error importing XLSX library {filename}: {e}")
            raise RuntimeError(f"Failed to import XLSX library: {e}") from e
    
    def _add_category_components_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add category components from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            category_ref = row[ExcelConstants.COMPONENTS_CATEGORY_REF] or ""
            if category_ref:
                if category_ref not in version.categories:
                    category_name = row[ExcelConstants.COMPONENTS_CATEGORY_NAME] or ""
                    category_uuid = row[ExcelConstants.COMPONENTS_CATEGORY_UUID] or ""
                    category_component = IRCategoryComponent(
                        uuid=category_uuid,
                        ref=category_ref,
                        name=category_name
                    )
                    version.categories[category_uuid] = category_component
    
    def _add_component_definitions_from_excel(self, lib: IRLibrary, sheet) -> None:
        """Add component definitions from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            component_ref = row[ExcelConstants.COMPONENTS_COMPONENT_REF] or ""
            if component_ref:
                component_name = row[ExcelConstants.COMPONENTS_COMPONENT_NAME] or ""
                component_desc = row[ExcelConstants.COMPONENTS_COMPONENT_DESC] or ""
                category_ref = row[ExcelConstants.COMPONENTS_CATEGORY_REF] or ""
                risk_patterns_str = row[ExcelConstants.COMPONENTS_RISK_PATTERNS] or ""
                visible = row[ExcelConstants.COMPONENTS_VISIBLE] or ""
                component_uuid = row[ExcelConstants.COMPONENTS_COMPONENT_UUID] or ""
                
                component_definition = IRComponentDefinition(
                    uuid=component_uuid,
                    ref=component_ref,
                    name=component_name,
                    desc=component_desc,
                    category_ref=category_ref,
                    visible=visible
                )
                
                if risk_patterns_str:
                    risk_patterns = [rp.strip() for rp in risk_patterns_str.split(",") if rp.strip()]
                    component_definition.risk_pattern_refs = risk_patterns
                
                lib.component_definitions[component_uuid] = component_definition
    
    def _add_supported_standards_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add supported standards from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            supported_standard_ref = row[ExcelConstants.SUPPORTED_STANDARD_REF] or ""
            if supported_standard_ref:
                supported_standard_name = row[ExcelConstants.SUPPORTED_STANDARD_NAME] or ""
                supported_standard_uuid = row[ExcelConstants.SUPPORTED_STANDARD_UUID] or ""
                supported_standard = IRSupportedStandard(
                    uuid=supported_standard_uuid,
                    ref=supported_standard_ref,
                    name=supported_standard_name
                )
                version.supported_standards[supported_standard_uuid] = supported_standard
    
    def _add_standards_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add standards from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            supported_standard_ref = row[0] or ""
            if supported_standard_ref:
                standard_ref = row[1] or ""
                standard_uuid = row[2] or ""
                standard = IRStandard(
                    uuid=standard_uuid,
                    supported_standard_ref=supported_standard_ref,
                    standard_ref=standard_ref
                )
                version.standards[standard_uuid] = standard
    
    def _add_rules_from_excel(self, lib: IRLibrary, sheet) -> None:
        """Add rules from Excel sheet"""
        logger.debug("Importing rules...")
        
        # Rules
        included_rules = set()
        rules_by_name = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rule_name = row[ExcelConstants.RULES_NAME] or ""
            if rule_name and rule_name not in included_rules:
                rule_module = row[ExcelConstants.RULES_MODULE] or ""
                rule_gui = row[ExcelConstants.RULES_GUI] or ""
                rule = IRRule(
                    name=rule_name,
                    module=rule_module,
                    generated_by_gui=rule_gui
                )
                lib.rules.append(rule)
                rules_by_name[rule_name] = rule
                included_rules.add(rule_name)
        
        # Conditions
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rule_name = self._get_rule_name(sheet, row[ExcelConstants.RULES_NAME])
            c_name = row[ExcelConstants.RULES_CONDITION_NAME] or ""
            if c_name and rule_name in rules_by_name:
                c_value = row[ExcelConstants.RULES_CONDITION_VALUE] or ""
                c_field = row[ExcelConstants.RULES_CONDITION_FIELD] or ""
                condition = IRRuleCondition(
                    name=c_name,
                    field=c_field,
                    value=c_value
                )
                rules_by_name[rule_name].conditions.append(condition)
        
        # Actions
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rule_name = self._get_rule_name(sheet, row[ExcelConstants.RULES_NAME])
            a_name = row[ExcelConstants.RULES_ACTION_NAME] or ""
            if a_name and rule_name in rules_by_name:
                a_value = row[ExcelConstants.RULES_ACTION_VALUE] or ""
                a_project = row[ExcelConstants.RULES_ACTION_PROJECT] or ""
                action = IRRuleAction(
                    name=a_name,
                    value=a_value,
                    project=a_project
                )
                rules_by_name[rule_name].actions.append(action)
        
        logger.debug("Importing rules finished")
    
    def _get_rule_name(self, sheet, rule_name) -> str:
        """Get rule name, handling merged cells"""
        if rule_name:
            return rule_name
        
        # Find the last non-empty rule name above this row
        for row_num in range(sheet.max_row, 0, -1):
            cell_value = sheet.cell(row_num, ExcelConstants.RULES_NAME + 1).value
            if cell_value:
                return cell_value
        
        return ""
    
    def _add_risk_patterns_from_excel(self, lib: IRLibrary, sheet) -> None:
        """Add risk patterns from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            working_rp_ref = row[0] or ""
            working_rp_name = row[1] or ""
            working_rp_desc = row[2] or ""
            working_rp_uuid = row[3] or ""
            
            if working_rp_ref:
                working_rp = IRRiskPattern(
                    ref=working_rp_ref,
                    name=working_rp_name,
                    desc=working_rp_desc,
                    uuid=working_rp_uuid
                )
                lib.risk_patterns[working_rp.uuid] = working_rp
        
        logger.info("Import risk patterns finished")
    
    def _add_usecases_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add use cases from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            working_uc_ref = row[0] or ""
            working_uc_name = row[1] or ""
            working_uc_desc = row[2] or ""
            working_uc_uuid = row[3] or ""
            
            if working_uc_ref:
                working_uc = IRUseCase(
                    ref=working_uc_ref,
                    name=working_uc_name,
                    desc=working_uc_desc
                )
                working_uc.uuid = working_uc_uuid
                version.usecases[working_uc.uuid] = working_uc
        
        logger.info("Import use cases finished")
    
    def _add_threats_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add threats from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            working_t_ref = row[0] or ""
            working_t_name = row[1] or ""
            working_t_desc = row[2] or ""
            working_t_conf = row[3] or ""
            working_t_int = row[4] or ""
            working_t_av = row[5] or ""
            working_t_ee = row[6] or ""
            refs_str = row[7] or ""
            mitre_str = row[8] or ""
            stride_str = row[9] or ""
            working_t_uuid = row[10] or ""
            
            if working_t_ref:
                working_th = IRThreat(
                    ref=working_t_ref,
                    name=working_t_name,
                    desc=working_t_desc
                )
                working_th.uuid = working_t_uuid
                
                # Add references
                if refs_str:
                    refs = [ref.strip() for ref in refs_str.split(ExcelConstants.SEPARATOR) if ref.strip()]
                    for ref in refs:
                        ref_parts = ref.split(":", 1)
                        if len(ref_parts) == 2:
                            ref_name, ref_url = ref_parts
                            for reference in version.references.values():
                                if reference.name == ref_name and reference.url == ref_url:
                                    ref_key = str(uuid.uuid4())
                                    working_th.references[ref_key] = reference.uuid
                                    break
                
                working_th.risk_rating = IRRiskRating(
                    confidentiality=working_t_conf,
                    integrity=working_t_int,
                    availability=working_t_av,
                    ease_of_exploitation=working_t_ee
                )
                
                # Import mitre if present
                if mitre_str:
                    mitre_items = [item.strip() for item in mitre_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_th.mitre = mitre_items
                
                # Import stride if present
                if stride_str:
                    stride_items = [item.strip() for item in stride_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_th.stride = stride_items
                
                version.threats[working_th.uuid] = working_th
        
        logger.info("Import threats finished")
    
    def _add_weaknesses_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add weaknesses from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            working_w_ref = row[0] or ""
            working_w_name = row[1] or ""
            working_w_desc = row[2] or ""
            working_w_impact = row[3] or ""
            working_w_test_steps = row[4] or ""
            refs_str = row[5] or ""
            working_w_uuid = row[6] or ""
            
            if working_w_ref:
                working_w = IRWeakness(
                    ref=working_w_ref,
                    name=working_w_name,
                    desc=working_w_desc
                )
                working_w.uuid = working_w_uuid
                working_w.impact = working_w_impact
                
                test = IRTest(steps=working_w_test_steps)
                
                # Add test references
                if refs_str:
                    refs = [ref.strip() for ref in refs_str.split(ExcelConstants.SEPARATOR) if ref.strip()]
                    for ref in refs:
                        ref_parts = ref.split(":", 1)
                        if len(ref_parts) == 2:
                            ref_name, ref_url = ref_parts
                            for reference in version.references.values():
                                if reference.name == ref_name and reference.url == ref_url:
                                    ref_key = str(uuid.uuid4())
                                    test.references[ref_key] = reference.uuid
                                    break
                
                working_w.test = test
                version.weaknesses[working_w.uuid] = working_w
        
        logger.info("Import weaknesses finished")
    
    def _add_controls_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add controls from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            working_c_ref = row[0] or ""
            working_c_name = row[1] or ""
            working_c_desc = row[2] or ""
            working_c_state = row[3] or ""
            working_c_cost = row[4] or ""
            refs_str = row[5] or ""
            test_steps = row[6] or ""
            test_refs_str = row[7] or ""
            standards_str = row[8] or ""
            implementations_str = row[9] or ""
            base_standard_str = row[10] or ""
            base_standard_section_str = row[11] or ""
            scope_str = row[12] or ""
            mitre_str = row[13] or ""
            working_c_uuid = row[14] or ""
            
            if working_c_ref:
                working_c = IRControl(
                    ref=working_c_ref,
                    name=working_c_name,
                    desc=working_c_desc,
                    state=working_c_state,
                    cost=working_c_cost
                )
                working_c.uuid = working_c_uuid
                
                # Add references
                if refs_str:
                    refs = [ref.strip() for ref in refs_str.split(ExcelConstants.SEPARATOR) if ref.strip()]
                    for ref in refs:
                        ref_parts = ref.split(":", 1)
                        if len(ref_parts) == 2:
                            ref_name, ref_url = ref_parts
                            for reference in version.references.values():
                                if reference.name == ref_name and reference.url == ref_url:
                                    ref_key = str(uuid.uuid4())
                                    working_c.references[ref_key] = reference.uuid
                                    break
                
                test = IRTest(steps=test_steps)
                
                # Add test references
                if test_refs_str:
                    test_refs = [ref.strip() for ref in test_refs_str.split(ExcelConstants.SEPARATOR) if ref.strip()]
                    for ref in test_refs:
                        ref_parts = ref.split(":", 1)
                        if len(ref_parts) == 2:
                            ref_name, ref_url = ref_parts
                            for reference in version.references.values():
                                if reference.name == ref_name and reference.url == ref_url:
                                    ref_key = str(uuid.uuid4())
                                    test.references[ref_key] = reference.uuid
                                    break
                
                working_c.test = test
                
                # Add standards
                if standards_str:
                    standards = [std.strip() for std in standards_str.split(ExcelConstants.SEPARATOR) if std.strip()]
                    for standard in standards:
                        standard_parts = standard.split(":", 1)
                        if len(standard_parts) == 2:
                            supported_standard_ref, standard_ref = standard_parts
                            for standard_obj in version.standards.values():
                                if (standard_obj.supported_standard_ref == supported_standard_ref and 
                                    standard_obj.standard_ref == standard_ref):
                                    standard_key = str(uuid.uuid4())
                                    working_c.standards[standard_key] = standard_obj.uuid
                                    break
                
                # Add implementations
                if implementations_str:
                    implementations = [imp.strip() for imp in implementations_str.split(ExcelConstants.SEPARATOR) if imp.strip()]
                    working_c.implementations = implementations
                
                # Import baseStandard if present
                if base_standard_str:
                    base_standard_items = [item.strip() for item in base_standard_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_c.base_standard = base_standard_items
                
                # Import baseStandardSection if present
                if base_standard_section_str:
                    base_standard_section_items = [item.strip() for item in base_standard_section_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_c.base_standard_section = base_standard_section_items
                
                # Import scope if present
                if scope_str:
                    scope_items = [item.strip() for item in scope_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_c.scope = scope_items
                
                # Import mitre if present
                if mitre_str:
                    mitre_items = [item.strip() for item in mitre_str.split(ExcelConstants.SEPARATOR) if item.strip()]
                    working_c.mitre = mitre_items
                
                version.controls[working_c.uuid] = working_c
        
        logger.info("Import controls finished")
    
    def _add_references_from_excel(self, version: ILEVersion, sheet) -> None:
        """Add references from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0] or ""
            url = row[1] or ""
            uuid_val = row[2] or ""
            
            if name and url:
                reference = IRReference(name=name, url=url)
                reference.uuid = uuid_val
                version.references[reference.uuid] = reference
        
        logger.info("Import references finished")
    
    def _add_relations_from_excel(self, lib: IRLibrary, sheet) -> None:
        """Add relations from Excel sheet"""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rp = row[0] or ""
            uc = row[1] or ""
            t = row[2] or ""
            w = row[3] or ""
            c = row[4] or ""
            m = row[5] or ""
            
            if rp:  # At least risk pattern should be present
                rel = IRRelation(
                    risk_pattern_uuid=rp,
                    usecase_uuid=uc,
                    threat_uuid=t,
                    weakness_uuid=w,
                    control_uuid=c,
                    mitigation=m
                )
                lib.relations[rel.uuid] = rel
        
        logger.info("Import relations finished")
